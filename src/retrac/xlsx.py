# Copyright (c) 2026 Martial Systems LLC
"""IDEM quarterly waste-received XLSX. All types pooled. Indiana origin only."""

from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from retrac.config import TON_COLS
from retrac.counties import key_name
from retrac.errors import FetchError


def _as_date(val: Any) -> date:
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    raise FetchError(f"bad start date {val!r}")


def _ton(val: Any) -> float:
    if val is None or val == "":
        return 0.0
    return float(val)


def parse_received(
    path: Path, *, counties: dict[str, dict[str, Any]]
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise FetchError("openpyxl is required to read the IDEM XLSX") from exc
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    idx = {name: i for i, name in enumerate(header)}
    need = ["Start Date", "ID Number", "Facility Name", "Origin State", "Origin County", *TON_COLS]
    missing = [c for c in need if c not in idx]
    if missing:
        raise FetchError(f"XLSX missing columns {missing}")
    buckets: dict[tuple[str, str, int, int], float] = defaultdict(float)
    names: dict[str, str] = {}
    type_tot: dict[str, float] = {c: 0.0 for c in TON_COLS}
    n_in = 0
    n_out = 0
    unknown: set[str] = set()
    for rec in rows_iter:
        if rec[idx["Origin State"]] != "Indiana":
            n_out += 1
            continue
        origin_raw = rec[idx["Origin County"]]
        if not origin_raw:
            raise FetchError("Indiana origin row with empty county")
        ok = key_name(str(origin_raw))
        if ok not in counties:
            unknown.add(str(origin_raw))
            continue
        fid = str(rec[idx["ID Number"]] or "").strip()
        if not fid:
            raise FetchError("Indiana origin row with empty facility ID")
        names[fid] = str(rec[idx["Facility Name"]] or fid).strip()
        start = _as_date(rec[idx["Start Date"]])
        q = (start.month - 1) // 3 + 1
        tons = 0.0
        for col in TON_COLS:
            t = _ton(rec[idx[col]])
            type_tot[col] += t
            tons += t
        buckets[(ok, fid, start.year, q)] += tons
        n_in += 1
    wb.close()
    if unknown:
        raise FetchError(f"unmatched origin counties {sorted(unknown)}")
    if n_in == 0:
        raise FetchError("no Indiana-origin rows in XLSX")
    out = []
    for (ok, fid, year, q), tons in sorted(buckets.items()):
        out.append(
            {
                "origin_key": ok,
                "origin_name": counties[ok]["name"],
                "facility_id": fid,
                "facility_name": names[fid],
                "year": int(year),
                "quarter": int(q),
                "tons": float(tons),
            }
        )
    stats = {
        "n_indiana_source_rows": n_in,
        "n_out_of_state": n_out,
        "type_totals_tons": {k: float(v) for k, v in type_tot.items()},
        "n_cells": len(out),
    }
    return out, stats
